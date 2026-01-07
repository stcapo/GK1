#!/usr/bin/env python3
"""
高考志愿系统数据ETL脚本
从上海和重庆的xlsx文件中提取数据，导入到MySQL数据库
"""

import os
import sys
import re
from pathlib import Path
from openpyxl import load_workbook
import pymysql

# 数据库配置
DB_CONFIG = {
    'host': 'localhost',
    'port': 3307,  # Docker映射端口
    'user': 'root',
    'password': '123456',
    'database': 'gaokao_system',
    'charset': 'utf8mb4'
}

# 数据目录
DATA_DIR = Path(__file__).parent.parent

def get_connection():
    """获取数据库连接"""
    return pymysql.connect(**DB_CONFIG)

def extract_university_from_name(full_name):
    """从专业组名称中提取大学名称"""
    # 常见格式: "上海交通大学(01)" 或 "复旦大学医学院(02)"
    match = re.match(r'^([^\(（]+)', full_name)
    if match:
        return match.group(1).strip()
    return full_name.strip()

def parse_shanghai_admission_2024(file_path):
    """解析上海2024年投档线数据"""
    print(f"解析文件: {file_path}")
    data = []
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            print(f"  跳过: 数据行数不足")
            return data
        
        # 尝试识别表头
        header = rows[0] if rows else []
        print(f"  表头: {header[:5]}...")
        
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            
            # 尝试提取: 院校名称, 专业组, 最低分, 位次
            university_name = None
            min_score = None
            min_rank = None
            
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                
                # 识别大学名称（包含"大学"或"学院"）
                if ('大学' in cell_str or '学院' in cell_str) and not university_name:
                    university_name = extract_university_from_name(cell_str)
                
                # 识别分数（3位数字，通常400-700之间）
                if isinstance(cell, (int, float)) and 300 <= cell <= 750:
                    if not min_score:
                        min_score = int(cell)
                    elif not min_rank and cell > 100:
                        min_rank = int(cell)
            
            if university_name and min_score:
                data.append({
                    'university': university_name,
                    'min_score': min_score,
                    'min_rank': min_rank,
                    'year': 2024,
                    'province': '上海'
                })
        
        wb.close()
        print(f"  提取了 {len(data)} 条记录")
        
    except Exception as e:
        print(f"  错误: {e}")
    
    return data

def insert_universities(conn, data):
    """插入大学数据（去重）"""
    cursor = conn.cursor()
    universities = set()
    
    for item in data:
        universities.add(item['university'])
    
    inserted = 0
    for univ in universities:
        # 检查是否已存在
        cursor.execute("SELECT id FROM universities WHERE name = %s", (univ,))
        if not cursor.fetchone():
            try:
                cursor.execute(
                    "INSERT INTO universities (name, province) VALUES (%s, %s)",
                    (univ, '上海')  # 默认上海，后续可根据数据更新
                )
                inserted += 1
            except Exception as e:
                print(f"  插入大学失败 {univ}: {e}")
    
    conn.commit()
    print(f"新增大学: {inserted} 所")
    return inserted

def insert_admission_data(conn, data):
    """插入录取数据"""
    cursor = conn.cursor()
    inserted = 0
    
    for item in data:
        # 获取大学ID
        cursor.execute("SELECT id FROM universities WHERE name = %s", (item['university'],))
        result = cursor.fetchone()
        if not result:
            continue
        
        university_id = result[0]
        
        try:
            cursor.execute("""
                INSERT INTO admission_data 
                (university_id, year, min_score, min_rank, source_province, batch)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE min_score = VALUES(min_score)
            """, (
                university_id,
                item['year'],
                item['min_score'],
                item.get('min_rank'),
                item['province'],
                '本科普通批'
            ))
            inserted += 1
        except Exception as e:
            print(f"  插入录取数据失败: {e}")
    
    conn.commit()
    print(f"新增录取数据: {inserted} 条")
    return inserted

def scan_and_import_shanghai():
    """扫描并导入上海数据"""
    shanghai_dir = DATA_DIR / '02-上海'
    admission_dir = shanghai_dir / '投档线'
    
    all_data = []
    
    # 扫描投档线目录
    if admission_dir.exists():
        for xlsx_file in admission_dir.rglob('*.xlsx'):
            # 优先处理普通批次数据
            if '普通' in xlsx_file.name or '平行志愿' in xlsx_file.name:
                data = parse_shanghai_admission_2024(xlsx_file)
                all_data.extend(data)
    
    if not all_data:
        print("未找到可导入的数据")
        return
    
    print(f"\n总共提取 {len(all_data)} 条记录")
    
    # 去重
    seen = set()
    unique_data = []
    for item in all_data:
        key = (item['university'], item['year'], item['min_score'])
        if key not in seen:
            seen.add(key)
            unique_data.append(item)
    
    print(f"去重后 {len(unique_data)} 条记录")
    
    # 导入数据库
    conn = get_connection()
    try:
        insert_universities(conn, unique_data)
        insert_admission_data(conn, unique_data)
    finally:
        conn.close()

def main():
    print("=" * 50)
    print("高考志愿系统 - 数据ETL")
    print("=" * 50)
    
    # 测试数据库连接
    try:
        conn = get_connection()
        print("✓ 数据库连接成功")
        conn.close()
    except Exception as e:
        print(f"✗ 数据库连接失败: {e}")
        print("请确保Docker容器正在运行")
        sys.exit(1)
    
    # 导入上海数据
    print("\n--- 导入上海数据 ---")
    scan_and_import_shanghai()
    
    print("\n完成!")

if __name__ == '__main__':
    main()
